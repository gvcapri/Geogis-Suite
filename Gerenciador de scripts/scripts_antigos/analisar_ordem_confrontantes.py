from __future__ import annotations

import argparse
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DOCX_PADRAO = Path(__file__).with_name("MD_LOTES_INTEGRAL_att - Copia.docx")
ORDEM_INTERNO = ("frente", "direita", "fundo", "esquerda")
ORDEM_EXTERNO = ("frente", "esquerda", "fundo", "direita")


def normalizar(texto: str) -> str:
    """Remove acentos e padroniza para facilitar a busca dos lados."""
    texto = texto.lower()
    return "".join(
        c for c in unicodedata.normalize("NFKD", texto)
        if not unicodedata.combining(c)
    )


def extrair_ordem_lados(descricao: str) -> tuple[list[str], list[str]]:
    """
    Retorna duas ordens:
    - ordem_completa: inclui repeticoes, util para lotes irregulares.
    - ordem_principal: primeira aparicao de cada lado.
    """
    texto = normalizar(descricao)
    padroes = [
        ("frente", r"\bde frente\b"),
        ("direita", r"\bdo lado direito\b"),
        ("fundo", r"\bao fundo\b"),
        ("esquerda", r"\bdo lado esquerdo\b"),
    ]

    ocorrencias: list[tuple[int, str]] = []
    for lado, padrao in padroes:
        for achado in re.finditer(padrao, texto):
            ocorrencias.append((achado.start(), lado))

    ordem_completa = [lado for _, lado in sorted(ocorrencias)]

    ordem_principal: list[str] = []
    for lado in ordem_completa:
        if lado not in ordem_principal:
            ordem_principal.append(lado)

    return ordem_completa, ordem_principal


def ler_memoriais(caminho_docx: Path) -> list[dict[str, str]]:
    doc = Document(caminho_docx)
    paragrafos = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    registros: list[dict[str, str]] = []

    for indice, texto in enumerate(paragrafos):
        marcador = re.search(
            r"PROPRIEDADE:\s*LOTE:\s*([^\n]+?)\s*-\s*QUADRA:\s*([^\n]+)",
            texto,
            flags=re.IGNORECASE,
        )
        if not marcador:
            continue

        lote = marcador.group(1).strip()
        quadra = marcador.group(2).strip()
        descricao = " ".join(paragrafos[indice + 1: indice + 4])
        ordem_completa, ordem_principal = extrair_ordem_lados(descricao)
        lado, observacao = classificar_lado(ordem_principal)

        registros.append({
            "quadra": quadra,
            "lote": lote,
            "lado": lado,
            "observacao": observacao,
            "_ordem_principal": " > ".join(ordem_principal),
            "_ordem_completa": " > ".join(ordem_completa),
        })

    return registros


def classificar_lado(ordem_principal: list[str]) -> tuple[str, str]:
    ordem = tuple(ordem_principal)
    if ordem == ORDEM_INTERNO:
        return "Interno", ""
    if ordem == ORDEM_EXTERNO:
        return "Externo", ""
    if ordem_e_subsequencia(ordem, ORDEM_INTERNO):
        return "Interno", "Sequencia parcial"
    if ordem_e_subsequencia(ordem, ORDEM_EXTERNO):
        return "Externo", "Sequencia parcial"
    return "Nao classificado", "Conferir"


def ordem_e_subsequencia(ordem_encontrada: tuple[str, ...], ordem_base: tuple[str, ...]) -> bool:
    if not ordem_encontrada:
        return False

    posicao = 0
    for lado in ordem_encontrada:
        try:
            posicao = ordem_base.index(lado, posicao) + 1
        except ValueError:
            return False
    return True


def ordenar_quadra(valor: str) -> tuple[int, int | str]:
    numero = re.search(r"\d+", valor)
    if numero:
        return (0, int(numero.group()))
    return (1, valor)


def salvar_xlsx(registros: list[dict[str, str]], caminho_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Analise"

    cabecalhos = ["Quadra", "Lote", "Lado", "Observacao"]
    ws.append(cabecalhos)

    for registro in sorted(
        registros,
        key=lambda r: (ordenar_quadra(r["quadra"]), ordenar_lote(r["lote"])),
    ):
        ws.append([
            registro["quadra"],
            registro["lote"],
            registro["lado"],
            registro["observacao"],
        ])

    fill = PatternFill("solid", fgColor="1F4E78")
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for coluna in ws.columns:
        largura = max(len(str(celula.value or "")) for celula in coluna) + 2
        ws.column_dimensions[get_column_letter(coluna[0].column)].width = min(largura, 24)

    resumo = wb.create_sheet("Resumo")
    resumo.append(["Lado", "Quantidade"])
    for lado, total in Counter(r["lado"] for r in registros).most_common():
        resumo.append([lado, total])
    for celula in resumo[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = fill
    resumo.column_dimensions["A"].width = 20
    resumo.column_dimensions["B"].width = 14

    conferir = wb.create_sheet("Conferir")
    conferir.append(["Quadra", "Lote", "Lado", "Observacao"])
    for registro in registros:
        if registro["observacao"]:
            conferir.append([
                registro["quadra"],
                registro["lote"],
                registro["lado"],
                registro["observacao"],
            ])
    for celula in conferir[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = fill
    conferir.column_dimensions["A"].width = 14
    conferir.column_dimensions["B"].width = 14
    conferir.column_dimensions["C"].width = 18
    conferir.column_dimensions["D"].width = 24

    wb.save(caminho_xlsx)


def ordenar_lote(valor: str) -> tuple[int, int | str]:
    numero = re.search(r"\d+", valor)
    if numero:
        return (0, int(numero.group()))
    return (1, valor)


def imprimir_resumo(registros: list[dict[str, str]], caminho_xlsx: Path) -> None:
    por_quadra: dict[str, list[dict[str, str]]] = defaultdict(list)
    for registro in registros:
        por_quadra[registro["quadra"]].append(registro)

    distribuicao = Counter(r["lado"] for r in registros)

    print("ANALISE DE ORDEM DOS CONFRONTANTES")
    print(f"Total de lotes analisados: {len(registros)}")
    print(f"Total de quadras/blocos: {len(por_quadra)}")
    print()

    print("Distribuicao por lado:")
    for lado, total in distribuicao.most_common():
        print(f"- {lado}: {total}")

    com_observacao = sum(1 for r in registros if r["observacao"])
    print(f"Lotes identificados para conferir: {com_observacao}")

    print()
    print(f"Planilha XLSX salva em: {caminho_xlsx}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Analisa a ordem Frente/Direita/Fundo/Esquerda nos memoriais de lotes."
    )
    parser.add_argument(
        "docx",
        nargs="?",
        type=Path,
        default=DOCX_PADRAO,
        help="Caminho do arquivo DOCX. Se omitido, usa o DOCX na mesma pasta do script.",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Caminho do XLSX de saida. Se omitido, cria ao lado do DOCX.",
    )
    args = parser.parse_args()

    caminho_docx = args.docx.expanduser().resolve()
    if not caminho_docx.exists():
        raise FileNotFoundError(f"Arquivo DOCX nao encontrado: {caminho_docx}")

    caminho_xlsx = args.xlsx
    if caminho_xlsx is None:
        caminho_xlsx = caminho_docx.with_name("analise_lado_confrontantes.xlsx")
    else:
        caminho_xlsx = caminho_xlsx.expanduser().resolve()

    registros = ler_memoriais(caminho_docx)
    salvar_xlsx(registros, caminho_xlsx)
    imprimir_resumo(registros, caminho_xlsx)


if __name__ == "__main__":
    main()
