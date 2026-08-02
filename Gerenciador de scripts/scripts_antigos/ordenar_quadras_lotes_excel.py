#!/usr/bin/env python3
"""Ordena quadras e lotes em arquivos Excel sem alterar os identificadores.

O programa pode ser usado com uma janela de seleção de arquivos ou pela linha de
comando. Somente abas com cabeçalhos reconhecidos de quadra e lote são
ordenadas; abas chamadas ``Controle`` e abas com mesclagens no corpo dos dados
são preservadas sem alteração.
"""

from __future__ import annotations

import argparse
import copy
import os
import re
import sys
import tempfile
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator, TranslatorError


QUADRA_ALIASES = {
    "Quadra",
    "QD",
    "Nº Quadra",
    "Número da Quadra",
    "Numero Quadra",
}
LOTE_ALIASES = {
    "Lote",
    "Lotes",
    "LT",
    "Nº Lote",
    "Número do Lote",
    "Numero Lote",
}
LINHAS_PARA_LOCALIZAR_CABECALHO = 30


def remover_acentos(texto: str) -> str:
    """Retorna texto sem marcas diacríticas, apenas para comparação."""
    normalizado = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in normalizado if not unicodedata.combining(c))


def normalizar_cabecalho(valor: Any) -> str:
    """Normaliza um cabeçalho ignorando caixa, acentos, espaços e sinais."""
    if valor is None:
        return ""
    texto = remover_acentos(str(valor)).casefold()
    return re.sub(r"[^a-z0-9]+", "", texto)


QUADRA_NORMALIZADA = {normalizar_cabecalho(v) for v in QUADRA_ALIASES}
LOTE_NORMALIZADO = {normalizar_cabecalho(v) for v in LOTE_ALIASES}


def _romano_para_inteiro(texto: str) -> int | None:
    """Converte um algarismo romano válido; retorna None se for inválido."""
    valores = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    if not texto or any(c not in valores for c in texto):
        return None
    total = 0
    anterior = 0
    for caractere in reversed(texto):
        atual = valores[caractere]
        if atual < anterior:
            total -= atual
        else:
            total += atual
            anterior = atual
    # Evita aceitar combinações romanas arbitrárias.
    milhares = "M" * (total // 1000)
    resto = total % 1000
    centenas = ("", "C", "CC", "CCC", "CD", "D", "DC", "DCC", "DCCC", "CM")
    dezenas = ("", "X", "XX", "XXX", "XL", "L", "LX", "LXX", "LXXX", "XC")
    unidades = ("", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX")
    canonico = milhares + centenas[resto // 100] + dezenas[(resto % 100) // 10] + unidades[resto % 10]
    return total if canonico == texto else None


def _converter_romano_final(texto: str) -> str:
    """Converte romano isolado no fim de séries como 'AV. III'."""
    correspondencia = re.fullmatch(r"(.*[A-Z])([\s.\-/]+)([IVXLCDM]+)", texto)
    if not correspondencia:
        return texto
    numero = _romano_para_inteiro(correspondencia.group(3))
    if numero is None:
        return texto
    return f"{correspondencia.group(1)} {numero}"


def chave_natural(valor: Any) -> tuple[Any, ...]:
    """Cria chave natural sem modificar o valor armazenado no Excel.

    Categorias: 0 = começa por número; 1 = somente letras; 2 = série textual
    com número; 3 = texto; 9 = vazio. Dentro da categoria, números são inteiros
    e textos são comparados sem acentos ou distinção de caixa. Hífens, barras,
    pontos e espaços delimitam tokens. A palavra ``E`` entre números também é
    tratada como separador de um agrupamento.
    """
    if valor is None:
        return (9, ())
    texto_original = str(valor)
    if not texto_original.strip():
        return (9, ())

    texto = remover_acentos(texto_original).upper().strip()
    texto = _converter_romano_final(texto)
    texto = re.sub(r"(?<=\d)\s+E\s+(?=\d)", " ", texto)

    if re.match(r"^\d", texto):
        categoria = 0
    elif re.fullmatch(r"[A-Z]+", re.sub(r"[^A-Z]", "", texto)) and not re.search(r"\d", texto):
        # Somente um bloco de letras, desconsiderando sinais/espaços.
        blocos = re.findall(r"[A-Z]+", texto)
        categoria = 1 if len(blocos) == 1 else 3
    elif re.search(r"\d", texto):
        categoria = 2
    else:
        categoria = 3

    tokens: list[tuple[int, Any]] = []
    for token in re.findall(r"\d+|[A-Z]+", texto):
        if token.isdigit():
            tokens.append((0, int(token)))
        else:
            tokens.append((1, token))
    return (categoria, tuple(tokens))


def localizar_cabecalho(ws: Any) -> tuple[int, int, int] | None:
    """Localiza linha, coluna de quadra e coluna de lote nas 30 primeiras linhas."""
    limite = min(ws.max_row, LINHAS_PARA_LOCALIZAR_CABECALHO)
    for numero_linha in range(1, limite + 1):
        coluna_quadra = None
        coluna_lote = None
        for celula in ws[numero_linha]:
            nome = normalizar_cabecalho(celula.value)
            if coluna_quadra is None and nome in QUADRA_NORMALIZADA:
                coluna_quadra = celula.column
            if coluna_lote is None and nome in LOTE_NORMALIZADO:
                coluna_lote = celula.column
        if coluna_quadra is not None and coluna_lote is not None:
            return numero_linha, coluna_quadra, coluna_lote
    return None


def _mesclagem_no_corpo(ws: Any, primeira_linha_dados: int) -> str | None:
    for intervalo in ws.merged_cells.ranges:
        if intervalo.max_row >= primeira_linha_dados:
            return str(intervalo)
    return None


def _traduzir_formula(valor: Any, origem: str, destino: str) -> Any:
    if not (isinstance(valor, str) and valor.startswith("=")) or origem == destino:
        return valor
    try:
        return Translator(valor, origin=origem).translate_formula(destino)
    except (TranslatorError, TypeError, ValueError):
        # Fórmulas especiais ou referências não traduzíveis são preservadas.
        return valor


@dataclass
class CelulaCopiada:
    valor: Any
    estilo: Any
    comentario: Any
    hiperlink: Any


@dataclass
class LinhaCopiada:
    numero_original: int
    celulas: list[CelulaCopiada]
    dimensao: Any
    chave_quadra: tuple[Any, ...]
    chave_lote: tuple[Any, ...]


@dataclass
class ResultadoAba:
    nome: str
    status: str
    detalhes: str
    linhas: int = 0


@dataclass
class ResultadoArquivo:
    origem: Path
    status: str
    destino: Path | None = None
    detalhes: str = ""
    abas: list[ResultadoAba] = field(default_factory=list)


def _capturar_linha(
    ws: Any,
    numero_linha: int,
    max_coluna: int,
    coluna_quadra: int,
    coluna_lote: int,
) -> LinhaCopiada:
    celulas = []
    for coluna in range(1, max_coluna + 1):
        celula: Cell = ws.cell(numero_linha, coluna)
        celulas.append(
            CelulaCopiada(
                valor=copy.copy(celula.value),
                estilo=copy.copy(celula._style),
                comentario=copy.copy(celula.comment),
                hiperlink=copy.copy(celula.hyperlink),
            )
        )
    return LinhaCopiada(
        numero_original=numero_linha,
        celulas=celulas,
        dimensao=copy.copy(ws.row_dimensions[numero_linha]),
        chave_quadra=chave_natural(ws.cell(numero_linha, coluna_quadra).value),
        chave_lote=chave_natural(ws.cell(numero_linha, coluna_lote).value),
    )


def _escrever_linha(ws: Any, linha: LinhaCopiada, destino: int) -> None:
    for indice, copiada in enumerate(linha.celulas, start=1):
        celula_destino: Cell = ws.cell(destino, indice)
        origem = ws.cell(linha.numero_original, indice).coordinate
        celula_destino.value = _traduzir_formula(copiada.valor, origem, celula_destino.coordinate)
        celula_destino._style = copy.copy(copiada.estilo)
        celula_destino.comment = copy.copy(copiada.comentario)
        celula_destino._hyperlink = copy.copy(copiada.hiperlink)

    dimensao = copy.copy(linha.dimensao)
    dimensao.index = destino
    ws.row_dimensions[destino] = dimensao


def _valor_assinavel(valor: Any) -> tuple[str, str]:
    return type(valor).__name__, repr(valor)


def _formula_canonica(valor: Any, coordenada: str, coluna: int) -> Any:
    if isinstance(valor, str) and valor.startswith("="):
        destino = f"{ws_coluna(coluna)}1"
        return _traduzir_formula(valor, coordenada, destino)
    return valor


def ws_coluna(numero: int) -> str:
    """Converte índice de coluna em letras sem depender de estado do workbook."""
    letras = ""
    while numero:
        numero, resto = divmod(numero - 1, 26)
        letras = chr(65 + resto) + letras
    return letras


def _assinatura_linha(ws: Any, linha: int, max_coluna: int) -> tuple[Any, ...]:
    assinatura = []
    for coluna in range(1, max_coluna + 1):
        celula = ws.cell(linha, coluna)
        valor = _formula_canonica(celula.value, celula.coordinate, coluna)
        assinatura.append(_valor_assinavel(valor))
    return tuple(assinatura)


def _validar_ordenacao(
    ws: Any,
    primeira_linha: int,
    ultima_linha: int,
    coluna_quadra: int,
    coluna_lote: int,
) -> None:
    chaves = [
        (
            chave_natural(ws.cell(linha, coluna_quadra).value),
            chave_natural(ws.cell(linha, coluna_lote).value),
        )
        for linha in range(primeira_linha, ultima_linha + 1)
    ]
    if chaves != sorted(chaves):
        raise ValueError("a conferência detectou dados fora da ordem natural")


def ordenar_aba(ws: Any, cabecalho: tuple[int, int, int]) -> ResultadoAba:
    """Ordena uma aba e executa validações de integridade em memória."""
    linha_cabecalho, coluna_quadra, coluna_lote = cabecalho
    primeira = linha_cabecalho + 1
    ultima = ws.max_row
    max_coluna = ws.max_column
    if primeira > ultima:
        return ResultadoAba(ws.title, "ignorada", "não há linhas após o cabeçalho")

    mesclagem = _mesclagem_no_corpo(ws, primeira)
    if mesclagem:
        return ResultadoAba(
            ws.title,
            "ignorada",
            f"mesclagem {mesclagem} alcança o corpo dos dados",
        )

    quantidade_antes = ultima - primeira + 1
    assinaturas_antes = Counter(
        _assinatura_linha(ws, linha, max_coluna) for linha in range(primeira, ultima + 1)
    )
    pares_antes = Counter(
        (
            _valor_assinavel(ws.cell(linha, coluna_quadra).value),
            _valor_assinavel(ws.cell(linha, coluna_lote).value),
        )
        for linha in range(primeira, ultima + 1)
    )

    linhas = [
        _capturar_linha(ws, linha, max_coluna, coluna_quadra, coluna_lote)
        for linha in range(primeira, ultima + 1)
    ]
    # O número original garante estabilidade de maneira explícita.
    linhas.sort(key=lambda item: (item.chave_quadra, item.chave_lote, item.numero_original))
    for destino, linha in enumerate(linhas, start=primeira):
        _escrever_linha(ws, linha, destino)

    if ws.max_row != ultima or ws.max_column != max_coluna:
        raise ValueError("a quantidade de linhas ou colunas foi alterada")
    quantidade_depois = ws.max_row - primeira + 1
    if quantidade_antes != quantidade_depois:
        raise ValueError("a quantidade de linhas de dados foi alterada")

    assinaturas_depois = Counter(
        _assinatura_linha(ws, linha, max_coluna) for linha in range(primeira, ultima + 1)
    )
    if assinaturas_antes != assinaturas_depois:
        raise ValueError("o conjunto de linhas mudou durante a ordenação")

    pares_depois = Counter(
        (
            _valor_assinavel(ws.cell(linha, coluna_quadra).value),
            _valor_assinavel(ws.cell(linha, coluna_lote).value),
        )
        for linha in range(primeira, ultima + 1)
    )
    if pares_antes != pares_depois:
        raise ValueError("o conjunto de pares quadra + lote foi alterado")
    _validar_ordenacao(ws, primeira, ultima, coluna_quadra, coluna_lote)
    return ResultadoAba(ws.title, "ordenada", "validações concluídas", quantidade_antes)


def _nome_destino(
    origem: Path,
    pasta_saida: str | os.PathLike[str] | None = None,
    nome_saida: str | None = None,
) -> Path:
    pasta = Path(pasta_saida).expanduser().resolve() if pasta_saida else origem.parent
    pasta.mkdir(parents=True, exist_ok=True)
    if nome_saida:
        nome_seguro = Path(nome_saida.strip()).name
        if not nome_seguro.casefold().endswith(".xlsx"):
            nome_seguro += ".xlsx"
        candidato = pasta / nome_seguro
    else:
        candidato = pasta / f"{origem.stem}_ORDENADO.xlsx"
    if not candidato.exists():
        return candidato
    instante = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = candidato.stem
    candidato = pasta / f"{base}_{instante}.xlsx"
    contador = 2
    while candidato.exists():
        candidato = pasta / f"{base}_{instante}_{contador}.xlsx"
        contador += 1
    return candidato


def processar_arquivo(
    caminho: str | os.PathLike[str],
    pasta_saida: str | os.PathLike[str] | None = None,
    nome_saida: str | None = None,
) -> ResultadoArquivo:
    """Processa um .xlsx com gravação temporária, reabertura e troca atômica."""
    origem = Path(caminho).expanduser().resolve()
    if origem.suffix.casefold() != ".xlsx":
        return ResultadoArquivo(origem, "ignorado", detalhes="a extensão não é .xlsx")
    if not origem.is_file():
        return ResultadoArquivo(origem, "erro", detalhes="arquivo não encontrado")

    temporario: Path | None = None
    workbook = None
    try:
        try:
            workbook = load_workbook(origem, data_only=False, keep_links=True, rich_text=True)
        except TypeError:  # Compatibilidade com openpyxl anterior a 3.1.
            workbook = load_workbook(origem, data_only=False, keep_links=True)

        resultados_abas: list[ResultadoAba] = []
        for ws in workbook.worksheets:
            if normalizar_cabecalho(ws.title) == "controle":
                resultados_abas.append(ResultadoAba(ws.title, "preservada", "aba auxiliar Controle"))
                continue
            cabecalho = localizar_cabecalho(ws)
            if cabecalho is None:
                resultados_abas.append(
                    ResultadoAba(ws.title, "preservada", "colunas de quadra e lote não localizadas")
                )
                continue
            resultados_abas.append(ordenar_aba(ws, cabecalho))

        if not any(item.status == "ordenada" for item in resultados_abas):
            workbook.close()
            return ResultadoArquivo(
                origem,
                "ignorado",
                detalhes="nenhuma aba pôde ser ordenada",
                abas=resultados_abas,
            )

        destino = _nome_destino(origem, pasta_saida, nome_saida)
        descritor, nome_temp = tempfile.mkstemp(
            prefix=f".{origem.stem}_ordenando_", suffix=".xlsx", dir=destino.parent
        )
        os.close(descritor)
        temporario = Path(nome_temp)
        workbook.save(temporario)
        workbook.close()
        workbook = None

        # A reabertura do temporário valida o pacote OOXML antes da publicação.
        conferencia = load_workbook(temporario, data_only=False, read_only=True, keep_links=True)
        conferencia.close()
        os.replace(temporario, destino)
        temporario = None

        # Validação final do nome publicado.
        conferencia_final = load_workbook(destino, data_only=False, read_only=True, keep_links=True)
        conferencia_final.close()
        return ResultadoArquivo(origem, "processado", destino, abas=resultados_abas)
    except Exception as erro:  # Cada arquivo falha isoladamente no processamento em lote.
        if workbook is not None:
            workbook.close()
        if temporario is not None and temporario.exists():
            temporario.unlink()
        return ResultadoArquivo(origem, "erro", detalhes=f"{type(erro).__name__}: {erro}")


def processar_arquivos(
    caminhos: Iterable[str | os.PathLike[str]],
    pasta_saida: str | os.PathLike[str] | None = None,
    nome_saida: str | None = None,
) -> list[ResultadoArquivo]:
    lista = list(caminhos)
    resultados = []
    for caminho in lista:
        nome_individual = nome_saida
        if nome_saida and len(lista) > 1:
            nome_individual = f"{Path(nome_saida).stem}_{Path(caminho).stem}.xlsx"
        resultados.append(processar_arquivo(caminho, pasta_saida, nome_individual))
    return resultados


def formatar_resumo(resultados: Sequence[ResultadoArquivo]) -> str:
    linhas = ["Resumo do processamento", "=" * 24]
    for resultado in resultados:
        linhas.append(f"\n[{resultado.status.upper()}] {resultado.origem}")
        if resultado.destino:
            linhas.append(f"  Saída: {resultado.destino}")
        if resultado.detalhes:
            linhas.append(f"  {resultado.detalhes}")
        for aba in resultado.abas:
            quantidade = f" ({aba.linhas} linhas)" if aba.linhas else ""
            linhas.append(f"  - {aba.nome}: {aba.status}{quantidade} — {aba.detalhes}")
    totais = Counter(resultado.status for resultado in resultados)
    linhas.append(
        "\nTotais: "
        f"{totais['processado']} processado(s), "
        f"{totais['ignorado']} ignorado(s), "
        f"{totais['erro']} erro(s)."
    )
    return "\n".join(linhas)


def selecionar_arquivos_gui() -> list[str]:
    """Abre a janela nativa de seleção múltipla do Windows."""
    import tkinter as tk
    from tkinter import filedialog

    raiz = tk.Tk()
    raiz.withdraw()
    raiz.update()
    arquivos = filedialog.askopenfilenames(
        parent=raiz,
        title="Selecione uma ou várias planilhas Excel",
        filetypes=[("Planilhas Excel", "*.xlsx")],
    )
    raiz.destroy()
    return list(arquivos)


def selecionar_saida_gui(arquivos: Sequence[str]) -> tuple[str, str | None] | None:
    """Permite escolher a pasta e, opcionalmente, um nome ou nome-base."""
    import tkinter as tk
    from tkinter import filedialog, simpledialog

    raiz = tk.Tk()
    raiz.withdraw()
    raiz.update()
    pasta = filedialog.askdirectory(
        parent=raiz,
        title="Selecione a pasta de saída",
        initialdir=str(Path(arquivos[0]).parent),
    )
    if not pasta:
        raiz.destroy()
        return None
    orientacao = (
        "Digite o nome do arquivo de saída (opcional)."
        if len(arquivos) == 1
        else "Digite um nome-base para os arquivos de saída (opcional)."
    )
    nome = simpledialog.askstring(
        "Nome da saída",
        orientacao + "\nDeixe vazio para usar o nome automático.",
        parent=raiz,
    )
    raiz.destroy()
    return pasta, nome.strip() if nome and nome.strip() else None


def mostrar_resumo_gui(texto: str, houve_erro: bool = False) -> None:
    import tkinter as tk
    from tkinter import messagebox

    raiz = tk.Tk()
    raiz.withdraw()
    if houve_erro:
        messagebox.showwarning("Ordenação concluída com avisos", texto, parent=raiz)
    else:
        messagebox.showinfo("Ordenação concluída", texto, parent=raiz)
    raiz.destroy()


def construir_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Ordena naturalmente quadras e lotes em uma ou várias planilhas .xlsx."
    )
    parser.add_argument(
        "arquivos",
        nargs="*",
        help="Arquivos .xlsx. Se omitidos, será aberta uma janela de seleção.",
    )
    parser.add_argument("--pasta-saida", help="Pasta onde os resultados serão gravados.")
    parser.add_argument(
        "--nome-saida",
        help="Nome do resultado; com vários arquivos, funciona como nome-base.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    argumentos = construir_parser().parse_args(argv)
    usando_gui = not argumentos.arquivos
    try:
        arquivos = argumentos.arquivos or selecionar_arquivos_gui()
    except Exception as erro:
        print(f"Não foi possível abrir a interface gráfica: {erro}", file=sys.stderr)
        print("Informe os arquivos pela linha de comando.", file=sys.stderr)
        return 2
    if not arquivos:
        print("Nenhum arquivo selecionado.")
        return 0

    pasta_saida = argumentos.pasta_saida
    nome_saida = argumentos.nome_saida
    if usando_gui:
        configuracao_saida = selecionar_saida_gui(arquivos)
        if configuracao_saida is None:
            print("Operação cancelada: nenhuma pasta de saída selecionada.")
            return 0
        pasta_saida, nome_saida = configuracao_saida

    resultados = processar_arquivos(arquivos, pasta_saida, nome_saida)
    resumo = formatar_resumo(resultados)
    print(resumo)
    houve_erro = any(resultado.status == "erro" for resultado in resultados)
    if usando_gui:
        mostrar_resumo_gui(resumo, houve_erro)
    return 1 if houve_erro else 0


if __name__ == "__main__":
    raise SystemExit(main())
